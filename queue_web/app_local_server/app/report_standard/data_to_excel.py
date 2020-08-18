def get_xls(path, matrix, sheet_name, excel_name, data_name):
    """data matrix write to excel and set styles"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    def create_excel(sheet_name):
        """use openpyxl to create excel to work on"""
        workbook = openpyxl.Workbook()
        # worksheet = workbook.create_sheet(title='sheet2', index=0)
        worksheet = workbook['Sheet']
        worksheet.title = sheet_name

        return workbook, worksheet

    def load_excel(excel_name, sheet_name):
        """load exist excel"""
        workbook = openpyxl.load_workbook(excel_name)
        print('all sheet name:', workbook.sheetnames)

        worksheet = workbook[sheet_name]
        print('opened excel:', excel_name, 'with sheet:', sheet_name)

        return workbook, worksheet

    def write_excel(matrix, worksheet, head_line, data_title, shift_right=3, shift_down=5):
        """what is it:
        1. write data matrix into excel, if excel exist, write it under the exist
        how it works:
        1. find exist max row index
        2. create title and style
        3. set head line and style
        4. traversal write data
        """
        exist_row_index = worksheet.max_row                     # get exist table's max row index
        print('current max row_index:', exist_row_index)
        shift_down = exist_row_index + shift_down               # shift new table under the exist table

        # data title
        worksheet.merge_cells(start_row=1+shift_down, start_column=1, end_row=1+shift_down+7, end_column=3)
        worksheet.cell(1+shift_down, 1, data_title)             # input data title into merged big cell

        title_index = 'A' + str(1+shift_down)
        worksheet[title_index].font = Font(size=16, italic=True, bold=True)         # change title font
        # head line
        for j in range(len(head_line)):
            worksheet.cell(1+shift_down, j+1+shift_right, head_line[j])              # input head line
            head_index = get_column_letter(j+1+shift_right) + str(1+shift_down)
            worksheet[head_index].fill = PatternFill("solid", fgColor="D1EEEE")     # change head line color
            worksheet[head_index].font = Font(size=12, bold=True)                   # change head line font
        # matrix to excel
        for j in range(len(matrix)):
            for i in range(len(matrix[j])):
                worksheet.cell(i+2+shift_down,  j+1+shift_right, str(matrix[j][i]))

    def auto_width(matrix, worksheet, head_line, shift_right=3):
        """auto adjust column width
        how it works:
        1. get every string length into wideness matrix
        2. compare whole column to find maximum length
        """
        col_width = [0 for i in matrix]                             # create column width array
        head_width = [0 for i in head_line]                         # head_line width matrix

        for i in range(len(matrix)):
            width_array = [len(data) for data in matrix[i]]         # width of this column
            if width_array:
                col_width[i] = max(width_array)                         # save max width of each column into col_width
                head_width[i] = len(head_line[i])                       # get heading width
           
            col_letter = get_column_letter(i + 1 + shift_right)     # get column sequence letter
            worksheet.column_dimensions[col_letter].width = max(col_width[i], head_width[i])*1.2 + 2    # modify

    # Main process begin
    Excel_output_path = path + "%s.xlsx" % excel_name
    excel, worksheet = create_excel(sheet_name)                                    # create excel

    head_line = ['', 'Volume(L/S)', 'Percentage(%)', 'Static Pressure', '(Pa)',
                 'Total Pressure', '(Pa)', 'Uniformity', '', 'Torque(N/m)']         # define head_line

    write_excel(matrix, worksheet, head_line, data_name)                           # write sheet
    auto_width(matrix, worksheet, head_line)                                       # adjust column width

    excel.save(filename=Excel_output_path)
    print('create new excel:', excel_name, 'create new sheet:', sheet_name)